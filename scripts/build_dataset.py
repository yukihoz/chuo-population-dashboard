#!/usr/bin/env python3
import argparse
import hashlib
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
import xlrd


BASE_URL = "https://www.city.chuo.lg.jp"
AGE_INDEX = BASE_URL + "/kusei/gaiyou/toukeidate/jinkou/kakutoshibetsu/index.html"
AREA_INDEX = BASE_URL + "/kusei/gaiyou/toukeidate/jinkou/choubetsu/index.html"


@dataclass
class SourceFile:
    category: str
    date: str
    url: str
    path: Path
    sha256: str


def fetch_url(url: str, *, retries: int = 3) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "chuo-pop-dashboard/0.1"})
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                return response.read()
        except urllib.error.URLError:
            if attempt == retries - 1:
                raise
            time.sleep(1 + attempt)
    raise RuntimeError(f"failed to fetch {url}")


def links_from_html(html: str) -> list[tuple[str, str]]:
    return re.findall(r'<a\s+[^>]*href="([^"]+)"[^>]*>(.*?)</a>', html, flags=re.I | re.S)


def visible_text(html: str) -> str:
    text = re.sub(r"<[^>]+>", "", html)
    return re.sub(r"\s+", "", text)


def label_to_year(label: str, href: str) -> int | None:
    for value in (label, href):
        match = re.search(r"(20\d{2})", value)
        if match:
            return int(match.group(1))
    label = visible_text(label)
    match = re.search(r"令和(\d+)年", label)
    if match:
        return 2018 + int(match.group(1))
    match = re.search(r"平成(\d+)年", label)
    if match:
        return 1988 + int(match.group(1))
    match = re.search(r"平成元年", label)
    if match:
        return 1989
    return None


def discover_year_pages(index_url: str, since_year: int) -> list[tuple[int, str]]:
    html = fetch_url(index_url).decode("utf-8", "ignore")
    pages: list[tuple[int, str]] = []
    for href, label in links_from_html(html):
        year = label_to_year(label, href)
        if not year or year < since_year:
            continue
        pages.append((year, urllib.parse.urljoin(BASE_URL, href)))
    return sorted(set(pages))


def date_from_file_url(url: str) -> str | None:
    name = Path(urllib.parse.urlparse(url).path).name
    match = re.search(r"(20\d{2})(\d{2})", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-01"
    match = re.search(r"(?<!\d)([1-3]\d)(0[1-9]|1[0-2])(?!\d)", name)
    if match:
        year = 1988 + int(match.group(1))
        return f"{year}-{match.group(2)}-01"
    return None


def date_from_label(label: str) -> str | None:
    text = visible_text(label)
    match = re.search(r"令和(\d+)年(\d{1,2})月1日", text)
    if match:
        return f"{2018 + int(match.group(1)):04d}-{int(match.group(2)):02d}-01"
    match = re.search(r"平成(\d+)年(\d{1,2})月1日", text)
    if match:
        return f"{1988 + int(match.group(1)):04d}-{int(match.group(2)):02d}-01"
    match = re.search(r"平成元年(\d{1,2})月1日", text)
    if match:
        return f"1989-{int(match.group(1)):02d}-01"
    match = re.search(r"(20\d{2})年(\d{1,2})月1日", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-01"
    return None


def discover_excel_links(index_url: str, category: str, since_year: int) -> list[tuple[str, str]]:
    links: list[tuple[str, str]] = []
    required_path = "kakutoshibetsu" if category == "age" else "choubetsu"
    for _, page_url in discover_year_pages(index_url, since_year):
        if required_path not in page_url:
            continue
        html = fetch_url(page_url).decode("utf-8", "ignore")
        for href, label in links_from_html(html):
            if not href.lower().endswith((".xls", ".xlsx")):
                continue
            url = urllib.parse.urljoin(BASE_URL, href)
            date = date_from_file_url(url) or date_from_label(label)
            if date:
                links.append((date, url))
    return sorted(set(links))


def download_sources(cache_dir: Path, since_year: int, limit: int | None = None) -> list[SourceFile]:
    sources: list[SourceFile] = []
    categories = [
        ("age", AGE_INDEX),
        ("area", AREA_INDEX),
    ]
    for category, index_url in categories:
        links = discover_excel_links(index_url, category, since_year)
        if limit:
            links = links[-limit:]
        for date, url in links:
            target = cache_dir / category / Path(urllib.parse.urlparse(url).path).name
            target.parent.mkdir(parents=True, exist_ok=True)
            if not target.exists():
                target.write_bytes(fetch_url(url))
            digest = hashlib.sha256(target.read_bytes()).hexdigest()
            sources.append(SourceFile(category, date, url, target, digest))
    return sources


def date_from_workbook(path: Path) -> str | None:
    ws = active_sheet(path)
    values = []
    for row in ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True):
        values.extend(str(value) for value in row if value is not None)
    return date_from_label(" ".join(values))


def sources_from_cache(cache_dir: Path, since_year: int, limit: int | None = None) -> list[SourceFile]:
    sources: list[SourceFile] = []
    for category in ("age", "area"):
        files = sorted((cache_dir / category).glob("*.xls*"))
        if limit:
            files = files[-limit:]
        for path in files:
            date = date_from_file_url(str(path)) or date_from_workbook(path)
            if not date or int(date[:4]) < since_year:
                continue
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            sources.append(SourceFile(category, date, path.resolve().as_uri(), path, digest))
    return sources


def as_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace(",", "").strip()
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    return None


class Cell:
    def __init__(self, value):
        self.value = value


class XlrdSheet:
    def __init__(self, sheet):
        self.sheet = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row, column):
        return Cell(self.sheet.cell_value(row - 1, column - 1))

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        end = max_row or self.max_row
        for row_idx in range(min_row - 1, end):
            yield tuple(self.sheet.cell_value(row_idx, col_idx) for col_idx in range(self.max_column))


def active_sheet(path: Path):
    if path.suffix.lower() == ".xls":
        return XlrdSheet(xlrd.open_workbook(path).sheet_by_index(0))
    return load_workbook(path, data_only=True, read_only=True).active


def normalize_name(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("　", " ").strip()
    return text.translate(str.maketrans("０１２３４５６７８９", "0123456789"))


def split_town_chome(name: str) -> tuple[str, int | None]:
    match = re.match(r"(.+?)([0-9]+)丁目$", name)
    if not match:
        return name, None
    return match.group(1), int(match.group(2))


def parse_area_new(ws, date: str) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = normalize_name(row[0])
        male, female, total, households = map(as_int, row[1:5])
        if not name or None in (male, female, total, households):
            continue
        town, chome = split_town_chome(name)
        rows.append(
            {
                "date": date,
                "name": name,
                "town": town,
                "chome": chome,
                "sourceOrder": len(rows),
                "households": households,
                "male": male,
                "female": female,
                "total": total,
            }
        )
    return rows


def parse_area_old(ws, date: str) -> list[dict]:
    rows = []
    blocks = [(0, 6), (6, 12), (12, 18)]
    current_towns = ["", "", ""]
    for row in ws.iter_rows(min_row=6, values_only=True):
        for block_index, (start, end) in enumerate(blocks):
            values = row[start:end]
            town_cell = normalize_name(values[0])
            chome_cell = values[1]
            households, total, male, female = map(as_int, values[2:6])
            if town_cell:
                current_towns[block_index] = town_cell
            if chome_cell == "計" or None in (households, total, male, female):
                continue
            chome = as_int(chome_cell)
            current_town = current_towns[block_index]
            if not current_town:
                continue
            if "計" in current_town or current_town == "区全体":
                continue
            if chome is None:
                if not town_cell:
                    continue
                name = current_town
            else:
                name = f"{current_town}{chome}丁目"
            rows.append(
                {
                    "date": date,
                    "name": name,
                    "town": current_town,
                    "chome": chome,
                    "sourceOrder": len(rows),
                    "households": households,
                    "male": male,
                    "female": female,
                    "total": total,
                }
            )
    return rows


def parse_area_file(source: SourceFile) -> list[dict]:
    ws = active_sheet(source.path)
    headers = [normalize_name(ws.cell(3, col).value) for col in range(1, 6)]
    if headers[:5] == ["名称", "男", "女", "計", "世帯数"]:
        return parse_area_new(ws, source.date)
    return parse_area_old(ws, source.date)


def parse_age_file(source: SourceFile) -> list[dict]:
    ws = active_sheet(source.path)
    rows = []
    for row in ws.iter_rows(values_only=True):
        age = as_int(row[0] if row else None)
        if age is None or age < 0 or age > 120:
            continue
        if (
            len(row) >= 11
            and as_int(row[2]) is not None
            and as_int(row[4]) is not None
            and as_int(row[5]) is not None
            and as_int(row[8]) is not None
            and as_int(row[10]) is not None
        ):
            total, male, female = map(as_int, row[2:5])
            jp_total, jp_male, jp_female = map(as_int, row[5:8])
            foreign_total, foreign_male, foreign_female = map(as_int, row[8:11])
        elif len(row) >= 10 and as_int(row[1]) is not None and as_int(row[9]) is not None and as_int(row[7]) is not None:
            jp_male, jp_female, jp_total = map(as_int, row[1:4])
            foreign_male, foreign_female, foreign_total = map(as_int, row[4:7])
            male, female, total = map(as_int, row[7:10])
        elif len(row) >= 5 and as_int(row[2]) is not None:
            total, male, female = map(as_int, row[2:5])
            jp_total, jp_male, jp_female = total, male, female
            foreign_total, foreign_male, foreign_female = 0, 0, 0
        else:
            continue
        rows.append(
            {
                "date": source.date,
                "age": age,
                "total": total,
                "male": male,
                "female": female,
                "japaneseTotal": jp_total,
                "japaneseMale": jp_male,
                "japaneseFemale": jp_female,
                "foreignTotal": foreign_total,
                "foreignMale": foreign_male,
                "foreignFemale": foreign_female,
            }
        )
    return rows


def aggregate_towns(area_rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    for row in area_rows:
        key = (row["date"], row["town"])
        item = grouped.setdefault(
            key,
            {
                "date": row["date"],
                "town": row["town"],
                "households": 0,
                "male": 0,
                "female": 0,
                "total": 0,
            },
        )
        for field in ("households", "male", "female", "total"):
            item[field] += row[field]
    return sorted(grouped.values(), key=lambda row: (row["date"], row["town"]))


def build_dataset(cache_dir: Path, output: Path, since_year: int, limit: int | None, from_cache: bool = False) -> None:
    sources = sources_from_cache(cache_dir, since_year, limit) if from_cache else download_sources(cache_dir, since_year, limit)
    area_rows: list[dict] = []
    age_rows: list[dict] = []
    failures: list[dict] = []
    for source in sources:
        try:
            if source.category == "area":
                area_rows.extend(parse_area_file(source))
            elif source.category == "age":
                age_rows.extend(parse_age_file(source))
        except Exception as exc:
            failures.append({"category": source.category, "date": source.date, "url": source.url, "error": str(exc)})

    dates = sorted({row["date"] for row in area_rows} | {row["date"] for row in age_rows})
    payload = {
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "dateRange": [dates[0], dates[-1]] if dates else [None, None],
        "source": {
            "baseUrl": BASE_URL,
            "sinceYear": since_year,
            "fileCount": len(sources),
            "failures": failures,
        },
        "areaChome": sorted(area_rows, key=lambda row: (row["date"], row.get("sourceOrder", 0), row["name"])),
        "areaTown": aggregate_towns(area_rows),
        "age": sorted(age_rows, key=lambda row: (row["date"], row["age"])),
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {output}")
    print(f"Area rows: {len(payload['areaChome'])} chome, {len(payload['areaTown'])} town")
    print(f"Age rows: {len(payload['age'])}")
    if failures:
        print(f"Failures: {len(failures)}", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--since-year", type=int, default=1998)
    parser.add_argument("--cache-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--output", type=Path, default=Path("data/chuo_population.json"))
    parser.add_argument("--limit-per-category", type=int)
    parser.add_argument("--from-cache", action="store_true")
    args = parser.parse_args()
    build_dataset(args.cache_dir, args.output, args.since_year, args.limit_per_category, args.from_cache)


if __name__ == "__main__":
    main()
